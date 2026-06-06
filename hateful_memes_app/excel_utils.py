import csv, os
import pandas as pd
from openpyxl import load_workbook

def safe_sheet_name(name: str) -> str:
    """Excel-Sheet-Namen: max 31 Zeichen, keine verbotenen Sonderzeichen."""
    for ch in ["×", "/", "\\", "[", "]", ":", "*", "?", "'"]:
        name = name.replace(ch, "x")
    return name[:31]

def write_sheet(filepath: str, sheet_name: str, df: pd.DataFrame) -> None:
    sheet_name = safe_sheet_name(sheet_name)
    os.makedirs(os.path.dirname(os.path.abspath(filepath)), exist_ok=True)
    mode = "a" if os.path.exists(filepath) else "w"
    kw = {"if_sheet_exists": "replace"} if mode == "a" else {}
    with pd.ExcelWriter(filepath, engine="openpyxl", mode=mode, **kw) as w:
        df.to_excel(w, sheet_name=sheet_name, index=False)

def read_sheet(filepath: str, sheet_name: str) -> pd.DataFrame:
    return pd.read_excel(filepath, sheet_name=sheet_name, engine="openpyxl")

def get_sheet_names(filepath: str) -> list[str]:
    if not os.path.exists(filepath):
        return []
    wb = load_workbook(filepath, read_only=True)
    names = wb.sheetnames
    wb.close()
    return names

def read_prompts(filepath: str, sheet: str) -> dict[str, str]:
    df = read_sheet(filepath, sheet)
    return dict(zip(df["Name"].astype(str), df["Prompt"].astype(str)))

def append_to_csv(row: dict, csv_path: str) -> None:
    """Hängt eine Zeile an CSV an. Erstellt Datei mit Header wenn nötig."""
    file_exists = os.path.exists(csv_path)
    if file_exists:
        # Feldnamen aus existierender Datei lesen um Konsistenz zu gewährleisten
        with open(csv_path, "r", newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            fieldnames = next(reader, list(row.keys()))
    else:
        fieldnames = list(row.keys())

    with open(csv_path, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        if not file_exists:
            writer.writeheader()
        writer.writerow(row)

def csv_to_excel_sheet(csv_path: str, excel_path: str, sheet_name: str) -> None:
    write_sheet(excel_path, sheet_name, pd.read_csv(csv_path, encoding="utf-8"))
